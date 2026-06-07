"""
Scraper - Consulta Estado NIT Bolivia
Portal: https://siat.impuestos.gob.bo/rnc/public/consultas-estado-nit

Uso:
    python consulta_nit_bolivia.py --nit 1234567890
    python consulta_nit_bolivia.py --archivo nits.txt
    python consulta_nit_bolivia.py --nit 1234567890 --exportar resultado.xlsx

Requisitos:
    pip install requests beautifulsoup4 lxml openpyxl
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import argparse
import csv
import sys
from datetime import datetime

# ─── Configuración ────────────────────────────────────────────────────────────

BASE_URL = "https://siat.impuestos.gob.bo/rnc"
CONSULTA_URL = f"{BASE_URL}/public/consultas-estado-nit"
API_URL = f"{BASE_URL}/public/consultas-estado-nit"   # ajustar si el endpoint difiere

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "es-BO,es;q=0.9",
    "Referer": CONSULTA_URL,
    "X-Requested-With": "XMLHttpRequest",
}

DELAY_ENTRE_CONSULTAS = 1.5   # segundos entre peticiones (evitar bloqueo)

# ─── Sesión ───────────────────────────────────────────────────────────────────

session = requests.Session()
session.headers.update(HEADERS)


def obtener_token_csrf() -> str | None:
    """
    Carga la página principal para capturar el token CSRF / cookie de sesión.
    Muchos portales Angular/React no usan CSRF tradicional; si no hay token
    devuelve None y se continúa sin él.
    """
    try:
        resp = session.get(CONSULTA_URL, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        # Buscar meta CSRF (Thymeleaf / Laravel / Django style)
        meta = soup.find("meta", {"name": "_csrf"}) or soup.find("meta", {"name": "csrf-token"})
        if meta:
            token = meta.get("content")
            session.headers["X-CSRF-Token"] = token
            return token

        # Buscar en input hidden
        inp = soup.find("input", {"name": "_token"}) or soup.find("input", {"name": "csrf_token"})
        if inp:
            return inp.get("value")

        return None
    except requests.RequestException as e:
        print(f"[AVISO] No se pudo cargar la página principal: {e}")
        return None


def consultar_nit(nit: str) -> dict:
    """
    Realiza la consulta de estado de un NIT.
    Retorna un diccionario con los datos extraídos o un error.
    """
    nit = str(nit).strip()

    # ── Intento 1: API REST/JSON (común en Angular/Vue) ────────────────────
    api_endpoints = [
        f"{BASE_URL}/public/consultas-estado-nit/buscar?nit={nit}",
        f"{BASE_URL}/public/buscarContribuyente?nit={nit}",
        f"{BASE_URL}/api/contribuyente?nit={nit}",
        f"{BASE_URL}/public/contribuyente/{nit}",
    ]

    for endpoint in api_endpoints:
        try:
            resp = session.get(endpoint, timeout=15)
            if resp.status_code == 200 and resp.headers.get("Content-Type", "").startswith("application/json"):
                data = resp.json()
                return _parsear_json(nit, data)
        except (requests.RequestException, json.JSONDecodeError):
            continue

    # ── Intento 2: Formulario POST clásico ─────────────────────────────────
    try:
        payload = {
            "nit": nit,
            "nro_nit": nit,
            "txtNit": nit,
        }
        resp = session.post(CONSULTA_URL, data=payload, timeout=15)
        resp.raise_for_status()

        if resp.headers.get("Content-Type", "").startswith("application/json"):
            return _parsear_json(nit, resp.json())

        return _parsear_html(nit, resp.text)

    except requests.RequestException as e:
        return {"nit": nit, "estado": "ERROR", "error": str(e)}


def _parsear_json(nit: str, data: dict | list) -> dict:
    """Normaliza la respuesta JSON del servidor."""
    if isinstance(data, list):
        data = data[0] if data else {}

    # Mapeo de posibles nombres de campo (el portal puede variar)
    campos_posibles = {
        "razon_social":  ["razonSocial", "razon_social", "nombre", "denominacion", "nombreRazonSocial"],
        "estado":        ["estadoContribuyente", "estado", "estadoNit", "situacion"],
        "tipo_contrib":  ["tipoContribuyente", "tipo", "categoriaContribuyente"],
        "domicilio":     ["domicilio", "direccion", "domicilioFiscal"],
        "actividad":     ["actividadEconomica", "actividad", "codigoActividad"],
        "fecha_inicio":  ["fechaInicioActividades", "fechaInicio", "fecha_inicio"],
        "departamento":  ["departamento", "dpto"],
        "municipio":     ["municipio"],
    }

    resultado = {"nit": nit, "consultado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    for campo, posibles in campos_posibles.items():
        for key in posibles:
            if key in data and data[key] not in (None, "", "null"):
                resultado[campo] = data[key]
                break
        else:
            resultado[campo] = data.get(campo, "N/D")

    # Si el servidor devolvió un mensaje de "no encontrado"
    if not resultado.get("razon_social") or resultado["razon_social"] == "N/D":
        resultado["estado"] = data.get("mensaje", data.get("message", "NO ENCONTRADO"))

    return resultado


def _parsear_html(nit: str, html: str) -> dict:
    """Parsea la respuesta HTML buscando tablas / divs con los datos."""
    soup = BeautifulSoup(html, "lxml")
    resultado = {"nit": nit, "consultado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

    # ── Buscar en tablas ────────────────────────────────────────────────────
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cols = [c.get_text(strip=True) for c in row.find_all(["th", "td"])]
            if len(cols) >= 2:
                clave = cols[0].lower().replace(" ", "_").replace(":", "")
                valor = cols[1]
                resultado[clave] = valor

    # ── Buscar campos etiquetados (dl/dt/dd) ────────────────────────────────
    for dl in soup.find_all("dl"):
        dts = dl.find_all("dt")
        dds = dl.find_all("dd")
        for dt, dd in zip(dts, dds):
            clave = dt.get_text(strip=True).lower().replace(" ", "_").replace(":", "")
            resultado[clave] = dd.get_text(strip=True)

    # ── Buscar divs con clase que contengan "nit" o "contribuyente" ─────────
    for div in soup.find_all("div", class_=lambda c: c and any(
            x in c.lower() for x in ["contribuyente", "resultado", "nit", "razon"])):
        texto = div.get_text(" | ", strip=True)
        if texto:
            resultado.setdefault("info_adicional", texto[:300])

    if len(resultado) <= 2:
        resultado["estado"] = "No se encontraron datos o el portal requiere JavaScript"

    return resultado


# ─── Exportar ─────────────────────────────────────────────────────────────────

def exportar_csv(resultados: list[dict], archivo: str):
    if not resultados:
        return
    campos = list(resultados[0].keys())
    with open(archivo, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(resultados)
    print(f"✅ CSV guardado: {archivo}")


def exportar_xlsx(resultados: list[dict], archivo: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[AVISO] openpyxl no instalado. Exportando como CSV.")
        exportar_csv(resultados, archivo.replace(".xlsx", ".csv"))
        return

    if not resultados:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consulta NIT Bolivia"

    encabezado_fill = PatternFill("solid", fgColor="003087")
    encabezado_font = Font(color="FFFFFF", bold=True)

    campos = list(resultados[0].keys())
    for col_idx, campo in enumerate(campos, 1):
        cell = ws.cell(row=1, column=col_idx, value=campo.upper().replace("_", " "))
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(resultados, 2):
        for col_idx, campo in enumerate(campos, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(item.get(campo, "")))

    # Ajustar anchos
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(archivo)
    print(f"✅ Excel guardado: {archivo}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Consulta estado de NIT en el portal SIAT del SIN Bolivia"
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--nit", help="Número de NIT a consultar")
    grupo.add_argument("--archivo", help="Archivo .txt con un NIT por línea")
    parser.add_argument("--exportar", help="Archivo de salida (.csv o .xlsx)", default=None)
    parser.add_argument("--delay", type=float, default=DELAY_ENTRE_CONSULTAS,
                        help="Segundos entre consultas (default: 1.5)")
    args = parser.parse_args()

    print("=" * 60)
    print("  Consulta Estado NIT - SIN Bolivia (SIAT)")
    print("=" * 60)

    # Cargar sesión / CSRF
    print("⏳ Iniciando sesión en el portal...")
    obtener_token_csrf()

    # Recopilar NITs
    nits = []
    if args.nit:
        nits = [args.nit]
    else:
        with open(args.archivo, encoding="utf-8") as f:
            nits = [l.strip() for l in f if l.strip()]

    print(f"📋 {len(nits)} NIT(s) a consultar\n")

    resultados = []
    for i, nit in enumerate(nits, 1):
        print(f"[{i}/{len(nits)}] Consultando NIT: {nit} ...", end=" ")
        resultado = consultar_nit(nit)
        resultados.append(resultado)

        estado = resultado.get("estado", resultado.get("razon_social", "?"))
        print(f"→ {estado}")

        if i < len(nits):
            time.sleep(args.delay)

    # Mostrar resumen en consola
    print("\n" + "─" * 60)
    print("RESULTADOS:")
    print("─" * 60)
    for r in resultados:
        for k, v in r.items():
            print(f"  {k:<25} {v}")
        print()

    # Exportar
    if args.exportar:
        if args.exportar.endswith(".xlsx"):
            exportar_xlsx(resultados, args.exportar)
        else:
            exportar_csv(resultados, args.exportar)
    elif len(resultados) > 1:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        exportar_csv(resultados, f"resultados_nit_{ts}.csv")


if __name__ == "__main__":
    main()
