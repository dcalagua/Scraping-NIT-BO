"""
Scraper (Playwright) - Consulta Estado NIT Bolivia
Portal: https://siat.impuestos.gob.bo/rnc/public/consultas-estado-nit

El portal es una SPA Angular protegida por un chequeo SSO de Keycloak.
En Chrome/Chromium ese chequeo expira ("Timeout when waiting for 3rd
party check iframe message") y la app redirige a www.impuestos.gob.bo,
dejando la página en blanco. En Firefox el chequeo se completa y el
formulario carga con normalidad, por lo que este script usa Firefox.

En vez de scrapear el HTML resultante, el script intercepta la respuesta
JSON que la propia página obtiene de su API interna
(siatrest.impuestos.gob.bo/.../rest/cons/estado-nit/{nit}), que es
exactamente lo que el formulario muestra en pantalla.

Uso:
    python consulta_nit_bolivia_playwright.py --nit 555162024
    python consulta_nit_bolivia_playwright.py --archivo nits.txt
    python consulta_nit_bolivia_playwright.py --nit 555162024 --exportar resultado.xlsx
    python consulta_nit_bolivia_playwright.py --nit 555162024 --headed   (para ver el navegador)

Requisitos:
    pip install playwright openpyxl
    python -m playwright install firefox
"""

import argparse
import asyncio
import csv
from datetime import datetime

from playwright.async_api import async_playwright, TimeoutError as PWTimeoutError

URL = "https://siat.impuestos.gob.bo/rnc/public/consultas-estado-nit"
DELAY_ENTRE_CONSULTAS = 1.5  # segundos entre consultas (evitar sobrecargar el portal)
TIMEOUT_RESPUESTA_MS = 25_000


# ─── Consulta individual ──────────────────────────────────────────────────────

async def consultar_nit(page, nit: str) -> dict:
    """
    Escribe el NIT en el formulario, presiona "Buscar" e intercepta la
    respuesta JSON de la API interna del portal. Retorna un diccionario
    normalizado con los datos del contribuyente o el motivo del error.
    """
    nit = str(nit).strip()
    resultado = {"nit": nit, "consultado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

    try:
        campo = await page.wait_for_selector("input[id^='mat-input']", timeout=TIMEOUT_RESPUESTA_MS)
        await campo.click()
        await campo.fill("")
        await campo.type(nit, delay=80)

        boton = await page.query_selector("button:has-text('Buscar')")

        async with page.expect_response(
            lambda r: "/rest/cons/estado-nit/" in r.url, timeout=TIMEOUT_RESPUESTA_MS
        ) as resp_info:
            await boton.click()
        resp = await resp_info.value

        if resp.status != 200:
            resultado["estado"] = f"ERROR HTTP {resp.status}"
            return resultado

        data = await resp.json()
    except PWTimeoutError:
        resultado["estado"] = "ERROR: tiempo de espera agotado (sin respuesta del portal)"
        return resultado
    except Exception as e:
        resultado["estado"] = f"ERROR: {e}"
        return resultado

    if not data.get("transaccion", False):
        mensajes = data.get("mensajes") or []
        descripcion = mensajes[0].get("descripcion") if mensajes else "NIT no encontrado"
        resultado["estado"] = descripcion
        return resultado

    datos = data.get("datosContribuyente", {})
    resultado.update({
        "razon_social":         datos.get("nombreRazonSocial", "N/D"),
        "estado":               datos.get("estadoContribuyente", "N/D"),
        "estado_actividad":     datos.get("estadoActividad", "N/D"),
        "tipo_contribuyente":   datos.get("tipoContribuyente", "N/D"),
        "regimen_contribuyente": datos.get("regimenContribuyente", "N/D"),
    })
    return resultado


# ─── Exportar ─────────────────────────────────────────────────────────────────

def exportar_csv(resultados: list, archivo: str):
    if not resultados:
        return
    campos = list(resultados[0].keys())
    with open(archivo, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(resultados)
    print(f"CSV guardado: {archivo}")


def exportar_xlsx(resultados: list, archivo: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[AVISO] openpyxl no instalado. Exportando como CSV.")
        exportar_csv(resultados, archivo.replace(".xlsx", ".csv"))
        return

    if not resultados:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consulta NIT Bolivia"

    encabezado_fill = PatternFill("solid", fgColor="003087")
    encabezado_font = Font(color="FFFFFF", bold=True)

    campos = list(resultados[0].keys())
    for col_idx, campo in enumerate(campos, 1):
        cell = ws.cell(row=1, column=col_idx, value=campo.upper().replace("_", " "))
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(resultados, 2):
        for col_idx, campo in enumerate(campos, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(item.get(campo, "")))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(archivo)
    print(f"Excel guardado: {archivo}")


# ─── Flujo principal ──────────────────────────────────────────────────────────

async def ejecutar(nits: list, headless: bool, delay: float) -> list:
    resultados = []
    async with async_playwright() as p:
        # Firefox: en Chrome/Chromium el chequeo SSO de Keycloak nunca
        # responde y la SPA redirige fuera de la página de consulta.
        browser = await p.firefox.launch(headless=headless)
        context = await browser.new_context(locale="es-BO")
        page = await context.new_page()

        print("Cargando el portal SIAT (puede tardar ~10s por el login SSO)...")
        await page.goto(URL, timeout=60_000)
        await page.wait_for_selector("input[id^='mat-input']", timeout=60_000)
        await page.wait_for_timeout(1000)

        for i, nit in enumerate(nits, 1):
            print(f"[{i}/{len(nits)}] Consultando NIT: {nit} ...", end=" ")
            resultado = await consultar_nit(page, nit)
            resultados.append(resultado)
            print(f"-> {resultado.get('razon_social', resultado.get('estado', '?'))}")

            if i < len(nits):
                await asyncio.sleep(delay)

        await browser.close()
    return resultados


def main():
    parser = argparse.ArgumentParser(
        description="Consulta estado de NIT en el portal SIAT del SIN Bolivia (vía Playwright)"
    )
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--nit", help="Número de NIT a consultar")
    grupo.add_argument("--archivo", help="Archivo .txt con un NIT por línea")
    parser.add_argument("--exportar", help="Archivo de salida (.csv o .xlsx)", default=None)
    parser.add_argument("--delay", type=float, default=DELAY_ENTRE_CONSULTAS,
                        help="Segundos entre consultas (default: 1.5)")
    parser.add_argument("--headed", action="store_true",
                        help="Mostrar la ventana del navegador (por defecto corre en segundo plano)")
    args = parser.parse_args()

    print("=" * 60)
    print("  Consulta Estado NIT - SIN Bolivia (SIAT) — Playwright")
    print("=" * 60)

    if args.nit:
        nits = [args.nit]
    else:
        with open(args.archivo, encoding="utf-8") as f:
            nits = [l.strip() for l in f if l.strip()]

    print(f"{len(nits)} NIT(s) a consultar\n")

    resultados = asyncio.run(ejecutar(nits, headless=not args.headed, delay=args.delay))

    print("\n" + "─" * 60)
    print("RESULTADOS:")
    print("─" * 60)
    for r in resultados:
        for k, v in r.items():
            print(f"  {k:<25} {v}")
        print()

    if args.exportar:
        if args.exportar.endswith(".xlsx"):
            exportar_xlsx(resultados, args.exportar)
        else:
            exportar_csv(resultados, args.exportar)
    elif len(resultados) > 1:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        exportar_csv(resultados, f"resultados_nit_{ts}.csv")


if __name__ == "__main__":
    main()
